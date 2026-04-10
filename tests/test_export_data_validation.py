import pytest
import openpyxl
from io import BytesIO
from services.asset_service import AssetService
from services.dict_service import DictService


@pytest.fixture
def db_session():
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker
    from models.base import Base

    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(bind=engine)
    Session = sessionmaker(bind=engine)
    session = Session()
    yield session
    session.close()


@pytest.fixture
def asset_service(db_session):
    return AssetService(db_session)


@pytest.fixture
def dict_service(db_session):
    return DictService(db_session)


def test_export_with_data_validation(asset_service, dict_service):
    dict_service.create_dict("unit", "单位")
    dict_service.create_dict_item("unit", "unit1", "单位1")
    dict_service.create_dict_item("unit", "unit2", "单位2")
    
    dict_service.create_dict("system", "系统")
    dict_service.create_dict_item("system", "sys1", "系统1")
    dict_service.create_dict_item("system", "sys2", "系统2")
    
    dict_service.create_dict("location", "位置")
    dict_service.create_dict_item("location", "loc1", "位置1")
    dict_service.create_dict_item("location", "loc2", "位置2")
    
    dict_service.create_dict("server_type", "服务器类型")
    dict_service.create_dict_item("server_type", "type1", "类型1")
    dict_service.create_dict_item("server_type", "type2", "类型2")
    
    asset_service.create(
        unit_name="unit1",
        system_name="sys1",
        username="admin",
        password="pass123",
        ip="192.168.1.1",
        port=22,
        location="loc1",
        server_type="type1"
    )
    
    file_data = asset_service.export_assets(include_password=False)
    workbook = openpyxl.load_workbook(BytesIO(file_data), data_only=False)
    worksheet = workbook.active
    
    assert worksheet.max_row == 2
    
    headers = [cell.value for cell in worksheet[1]]
    print(f"Headers: {headers}")
    
    assert worksheet.cell(row=2, column=1).value == "单位1"
    assert worksheet.cell(row=2, column=2).value == "系统1"
    assert worksheet.cell(row=2, column=9).value == "位置1"
    assert worksheet.cell(row=2, column=10).value == "类型1"


def test_export_with_password_and_data_validation(asset_service, dict_service):
    dict_service.create_dict("unit", "单位")
    dict_service.create_dict_item("unit", "unit1", "单位1")
    
    dict_service.create_dict("system", "系统")
    dict_service.create_dict_item("system", "sys1", "系统1")
    
    dict_service.create_dict("location", "位置")
    dict_service.create_dict_item("location", "loc1", "位置1")
    
    dict_service.create_dict("server_type", "服务器类型")
    dict_service.create_dict_item("server_type", "type1", "类型1")
    
    asset_service.create(
        unit_name="unit1",
        system_name="sys1",
        username="admin",
        password="secret123",
        ip="192.168.1.1",
        port=22,
        location="loc1",
        server_type="type1"
    )
    
    file_data = asset_service.export_assets(include_password=True)
    workbook = openpyxl.load_workbook(BytesIO(file_data), data_only=False)
    worksheet = workbook.active
    
    assert worksheet.max_row == 2
    
    headers = [cell.value for cell in worksheet[1]]
    print(f"Headers with password: {headers}")
    assert "密码" in headers
    
    assert worksheet.cell(row=2, column=1).value == "单位1"
    assert worksheet.cell(row=2, column=2).value == "系统1"
    assert worksheet.cell(row=2, column=10).value == "位置1"
    assert worksheet.cell(row=2, column=11).value == "类型1"


def test_export_empty_with_data_validation(asset_service, dict_service):
    dict_service.create_dict("unit", "单位")
    dict_service.create_dict_item("unit", "unit1", "单位1")
    
    dict_service.create_dict("system", "系统")
    dict_service.create_dict_item("system", "sys1", "系统1")
    
    dict_service.create_dict("location", "位置")
    dict_service.create_dict_item("location", "loc1", "位置1")
    
    dict_service.create_dict("server_type", "服务器类型")
    dict_service.create_dict_item("server_type", "type1", "类型1")
    
    file_data = asset_service.export_assets(include_password=False)
    workbook = openpyxl.load_workbook(BytesIO(file_data), data_only=False)
    worksheet = workbook.active
    
    assert worksheet.max_row == 1
    
    headers = [cell.value for cell in worksheet[1]]
    assert "单位名称*" in headers
    assert "系统名称*" in headers
    assert "位置" in headers
    assert "服务器类型" in headers
