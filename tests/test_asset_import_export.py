import pytest
import openpyxl
import io
from io import BytesIO
from services.asset_service import AssetService
from services.dict_service import DictService


@pytest.fixture
def asset_service(db_session):
    return AssetService()


@pytest.fixture
def dict_service(db_session):
    return DictService()


def test_export_assets(asset_service):
    asset_service.create(
        unit_name="Unit1",
        system_name="System1",
        username="admin",
        password="pass123",
        ip="192.168.1.1",
        port=22
    )
    asset_service.create(
        unit_name="Unit2",
        system_name="System2",
        username="root",
        password="pass456",
        ip="192.168.1.2",
        port=22
    )

    file_data = asset_service.export_assets(include_password=False)
    assert file_data is not None
    assert len(file_data) > 0

    workbook = openpyxl.load_workbook(io.BytesIO(file_data))
    worksheet = workbook.active
    
    assert worksheet.max_row == 3
    
    headers = [cell.value for cell in worksheet[1]]
    expected_headers = [
        "单位名称", "系统名称", "IP地址", "IPv6地址", "用户名",
        "端口", "主机名", "业务服务", "位置", "服务器类型", "VIP", "备注"
    ]
    assert headers == expected_headers
    assert "密码" not in headers


def test_export_assets_with_password(asset_service):
    asset_service.create(
        unit_name="Unit1",
        system_name="System1",
        username="admin",
        password="secret123",
        ip="192.168.1.1",
        port=22
    )

    file_data = asset_service.export_assets(include_password=True)
    assert file_data is not None

    workbook = openpyxl.load_workbook(io.BytesIO(file_data))
    worksheet = workbook.active
    
    headers = [cell.value for cell in worksheet[1]]
    expected_headers = [
        "单位名称", "系统名称", "IP地址", "IPv6地址", "用户名",
        "密码", "端口", "主机名", "业务服务", "位置", "服务器类型", "VIP", "备注"
    ]
    assert headers == expected_headers
    assert "密码" in headers
    
    password_col = headers.index("密码") + 1
    assert worksheet.cell(row=2, column=password_col).value == "secret123"


def test_export_selected_assets(asset_service):
    asset_id1 = asset_service.create("Unit1", "System1", "admin", "pass", ip="192.168.1.1", port=22)
    asset_id2 = asset_service.create("Unit2", "System2", "admin", "pass", ip="192.168.1.2", port=22)
    asset_service.create("Unit3", "System3", "admin", "pass", ip="192.168.1.3", port=22)

    file_data = asset_service.export_assets(asset_ids=[asset_id1, asset_id2])
    workbook = openpyxl.load_workbook(io.BytesIO(file_data))
    worksheet = workbook.active
    
    assert worksheet.max_row == 3


def test_import_assets_excel(asset_service):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.append([
        "单位名称", "系统名称", "IP地址", "IPv6地址", "用户名", "密码", "端口"
    ])
    worksheet.append(["Unit1", "System1", "192.168.1.1", "", "admin", "pass123", 22])
    worksheet.append(["Unit2", "System2", "192.168.1.2", "", "root", "pass456", 22])
    
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    success_count, fail_count, errors = asset_service.import_assets(
        file_data=output.getvalue()
    )

    assert success_count == 2
    assert fail_count == 0
    assert len(errors) == 0

    assets = asset_service.get_all()
    assert len(assets) == 2


def test_import_assets_missing_required_fields(asset_service):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.append([
        "单位名称", "系统名称", "用户名"
    ])
    worksheet.append(["Unit1", "", "admin"])

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    success_count, fail_count, errors = asset_service.import_assets(
        file_data=output.getvalue(),
        skip_errors=True
    )

    assert success_count == 0
    assert fail_count == 1
    assert len(errors) > 0
    assert "缺少必填字段" in errors[0]


def test_import_assets_update_existing(asset_service):
    asset_service.create("Unit1", "System1", "admin", "oldpass", ip="192.168.1.1", port=22)

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.append([
        "单位名称", "系统名称", "IP地址", "IPv6地址", "用户名", "密码", "端口"
    ])
    worksheet.append(["Unit1", "System1", "192.168.1.1", "", "admin", "newpass", 2222])

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    success_count, fail_count, errors = asset_service.import_assets(
        file_data=output.getvalue(),
        update_existing=True
    )

    assert success_count == 1
    assert fail_count == 0

    assets = asset_service.get_all()
    assert len(assets) == 1
    assert assets[0].username == "admin"
    assert assets[0].ip == "192.168.1.1"
    assert assets[0].port == 2222
    assert asset_service.get_password(assets[0].id) == "newpass"


def test_import_assets_duplicate_without_update(asset_service):
    asset_service.create("Unit1", "System1", "admin", "pass", ip="192.168.1.1", port=22)

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.append([
        "单位名称", "系统名称", "IP地址", "IPv6地址", "用户名", "密码"
    ])
    worksheet.append(["Unit1", "System1", "192.168.1.1", "", "admin", "pass"])

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    success_count, fail_count, errors = asset_service.import_assets(
        file_data=output.getvalue(),
        update_existing=False,
        skip_errors=True
    )

    assert success_count == 0
    assert fail_count == 1
    assert "资产已存在" in errors[0]


def test_import_assets_missing_ip_and_ipv6(asset_service):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.append([
        "单位名称", "系统名称", "IP地址", "IPv6地址", "用户名", "密码"
    ])
    worksheet.append(["Unit1", "System1", "", "", "admin", "pass"])

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    success_count, fail_count, errors = asset_service.import_assets(
        file_data=output.getvalue(),
        skip_errors=True
    )

    assert success_count == 0
    assert fail_count == 1
    assert "IP地址和IPv6地址至少需要填写一个" in errors[0]


def test_import_assets_same_unit_system_different_ip(asset_service):
    asset_service.create("Unit1", "System1", "admin", "pass", ip="192.168.1.1", port=22)

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.append([
        "单位名称", "系统名称", "IP地址", "IPv6地址", "用户名", "密码"
    ])
    worksheet.append(["Unit1", "System1", "192.168.1.2", "", "admin", "pass"])

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    success_count, fail_count, errors = asset_service.import_assets(
        file_data=output.getvalue()
    )

    assert success_count == 1
    assert fail_count == 0

    assets = asset_service.get_all()
    assert len(assets) == 2


def test_import_assets_same_unit_system_ip_different_user(asset_service):
    asset_service.create("Unit1", "System1", "admin", "pass", ip="192.168.1.1", port=22)

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.append([
        "单位名称", "系统名称", "IP地址", "IPv6地址", "用户名", "密码"
    ])
    worksheet.append(["Unit1", "System1", "192.168.1.1", "", "root", "pass"])

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    success_count, fail_count, errors = asset_service.import_assets(
        file_data=output.getvalue()
    )

    assert success_count == 1
    assert fail_count == 0

    assets = asset_service.get_all()
    assert len(assets) == 2


def test_import_assets_with_ipv6_only(asset_service):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.append([
        "单位名称", "系统名称", "IP地址", "IPv6地址", "用户名", "密码"
    ])
    worksheet.append(["Unit1", "System1", "", "2001:db8::1", "admin", "pass"])

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    success_count, fail_count, errors = asset_service.import_assets(
        file_data=output.getvalue()
    )

    assert success_count == 1
    assert fail_count == 0

    assets = asset_service.get_all()
    assert len(assets) == 1
    assert assets[0].ipv6 == "2001:db8::1"


def test_import_assets_skip_errors(asset_service):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.append([
        "单位名称", "系统名称", "IP地址", "IPv6地址", "用户名", "密码"
    ])
    worksheet.append(["Unit1", "System1", "192.168.1.1", "", "admin", "pass1"])
    worksheet.append(["", "System2", "192.168.1.2", "", "admin", "pass2"])
    worksheet.append(["Unit2", "System2", "192.168.1.3", "", "root", "pass3"])

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    success_count, fail_count, errors = asset_service.import_assets(
        file_data=output.getvalue(),
        skip_errors=True
    )

    assert success_count == 2
    assert fail_count == 1
    assert len(errors) == 1


def test_import_export_roundtrip(asset_service):
    asset_service.create("Unit1", "System1", "admin", "pass123", ip="192.168.1.1", port=22)
    asset_service.create("Unit2", "System2", "root", "pass456", ip="192.168.1.2", port=2222)

    file_data = asset_service.export_assets(include_password=True)

    from models.server_asset import ServerAsset
    from core.database import get_db
    with get_db() as db:
        db.query(ServerAsset).delete()
        db.commit()

    success_count, fail_count, errors = asset_service.import_assets(
        file_data=file_data
    )

    assert success_count == 2
    assert fail_count == 0

    assets = asset_service.get_all()
    assert len(assets) == 2
    assert assets[0].unit_name == "Unit1"
    assert assets[1].unit_name == "Unit2"


def test_export_empty_with_headers(asset_service):
    file_data = asset_service.export_assets(include_password=False)
    workbook = openpyxl.load_workbook(BytesIO(file_data))
    worksheet = workbook.active

    expected_columns = [
        "单位名称", "系统名称", "IP地址", "IPv6地址", "用户名",
        "端口", "主机名", "业务服务", "位置", "服务器类型", "VIP", "备注"
    ]

    headers = [cell.value for cell in worksheet[1]]
    assert headers == expected_columns
    assert worksheet.max_row == 1


def test_export_with_data_validation(asset_service, dict_service):
    dict_service.create_dict("unit", "单位")
    dict_service.create_dict_item("unit", "unit1", "单位1")
    dict_service.create_dict_item("unit", "unit2", "单位2")

    dict_service.create_dict("system", "系统")
    dict_service.create_dict_item("system", "sys1", "系统1")
    dict_service.create_dict_item("system", "sys2", "系统2")

    dict_service.create_dict("location", "位置")
    dict_service.create_dict_item("location", "loc1", "位置1")
    dict_service.create_dict_item("location", "loc2", "位置2")

    dict_service.create_dict("server_type", "服务器类型")
    dict_service.create_dict_item("server_type", "type1", "类型1")
    dict_service.create_dict_item("server_type", "type2", "类型2")

    asset_service.create(
        unit_name="unit1",
        system_name="sys1",
        username="admin",
        password="pass123",
        ip="192.168.1.1",
        port=22,
        location="loc1",
        server_type="type1"
    )

    file_data = asset_service.export_assets(include_password=False)
    workbook = openpyxl.load_workbook(BytesIO(file_data), data_only=False)
    worksheet = workbook.active

    assert worksheet.max_row == 2
    assert worksheet.cell(row=2, column=1).value == "单位1"
    assert worksheet.cell(row=2, column=2).value == "系统1"
    assert worksheet.cell(row=2, column=9).value == "位置1"
    assert worksheet.cell(row=2, column=10).value == "类型1"


def test_export_with_password_and_data_validation(asset_service, dict_service):
    dict_service.create_dict("unit", "单位")
    dict_service.create_dict_item("unit", "unit1", "单位1")

    dict_service.create_dict("system", "系统")
    dict_service.create_dict_item("system", "sys1", "系统1")

    dict_service.create_dict("location", "位置")
    dict_service.create_dict_item("location", "loc1", "位置1")

    dict_service.create_dict("server_type", "服务器类型")
    dict_service.create_dict_item("server_type", "type1", "类型1")

    asset_service.create(
        unit_name="unit1",
        system_name="sys1",
        username="admin",
        password="secret123",
        ip="192.168.1.1",
        port=22,
        location="loc1",
        server_type="type1"
    )

    file_data = asset_service.export_assets(include_password=True)
    workbook = openpyxl.load_workbook(BytesIO(file_data), data_only=False)
    worksheet = workbook.active

    assert worksheet.max_row == 2

    headers = [cell.value for cell in worksheet[1]]
    assert "密码" in headers

    assert worksheet.cell(row=2, column=1).value == "单位1"
    assert worksheet.cell(row=2, column=2).value == "系统1"
    assert worksheet.cell(row=2, column=10).value == "位置1"
    assert worksheet.cell(row=2, column=11).value == "类型1"


def test_export_empty_with_data_validation(asset_service, dict_service):
    dict_service.create_dict("unit", "单位")
    dict_service.create_dict_item("unit", "unit1", "单位1")

    dict_service.create_dict("system", "系统")
    dict_service.create_dict_item("system", "sys1", "系统1")

    dict_service.create_dict("location", "位置")
    dict_service.create_dict_item("location", "loc1", "位置1")

    dict_service.create_dict("server_type", "服务器类型")
    dict_service.create_dict_item("server_type", "type1", "类型1")

    file_data = asset_service.export_assets(include_password=False)
    workbook = openpyxl.load_workbook(BytesIO(file_data), data_only=False)
    worksheet = workbook.active

    assert worksheet.max_row == 1

    headers = [cell.value for cell in worksheet[1]]
    assert "单位名称" in headers
    assert "系统名称" in headers
    assert "位置" in headers
    assert "服务器类型" in headers
