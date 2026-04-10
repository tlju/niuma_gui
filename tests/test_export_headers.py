import pytest
import openpyxl
from io import BytesIO
from services.asset_service import AssetService


@pytest.fixture
def db_session():
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker
    from models.base import Base

    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(bind=engine)
    Session = sessionmaker(bind=engine)
    session = Session()
    yield session
    session.close()


@pytest.fixture
def asset_service(db_session):
    return AssetService(db_session)


def test_export_with_headers(asset_service):
    asset_service.create(
        unit_name="Unit1",
        system_name="System1",
        username="admin",
        password="pass123",
        ip="192.168.1.1",
        port=22
    )

    file_data = asset_service.export_assets(include_password=False)
    workbook = openpyxl.load_workbook(BytesIO(file_data))
    worksheet = workbook.active
    
    expected_columns = [
        "单位名称", "系统名称", "IP地址", "IPv6地址", "用户名",
        "端口", "主机名", "业务服务", "位置", "服务器类型", "VIP", "备注"
    ]
    
    headers = [cell.value for cell in worksheet[1]]
    assert headers == expected_columns
    assert worksheet.max_row == 2
    assert worksheet.cell(row=2, column=1).value == "Unit1"
    assert worksheet.cell(row=2, column=2).value == "System1"


def test_export_empty_with_headers(asset_service):
    file_data = asset_service.export_assets(include_password=False)
    workbook = openpyxl.load_workbook(BytesIO(file_data))
    worksheet = workbook.active
    
    expected_columns = [
        "单位名称", "系统名称", "IP地址", "IPv6地址", "用户名",
        "端口", "主机名", "业务服务", "位置", "服务器类型", "VIP", "备注"
    ]
    
    headers = [cell.value for cell in worksheet[1]]
    assert headers == expected_columns
    assert worksheet.max_row == 1


def test_export_with_password_column(asset_service):
    asset_service.create(
        unit_name="Unit1",
        system_name="System1",
        username="admin",
        password="secret123",
        ip="192.168.1.1",
        port=22
    )

    file_data = asset_service.export_assets(include_password=True)
    workbook = openpyxl.load_workbook(BytesIO(file_data))
    worksheet = workbook.active
    
    expected_columns = [
        "单位名称", "系统名称", "IP地址", "IPv6地址", "用户名",
        "密码", "端口", "主机名", "业务服务", "位置", "服务器类型", "VIP", "备注"
    ]
    
    headers = [cell.value for cell in worksheet[1]]
    assert headers == expected_columns
    assert "密码" in headers
    password_col = headers.index("密码") + 1
    assert worksheet.cell(row=2, column=password_col).value == "secret123"
