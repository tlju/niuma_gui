from sqlalchemy.orm import Session
from models.server_asset import ServerAsset
from services.crypto import CryptoManager
from services.dict_service import DictService
from services.audit_mixin import AuditMixin
from typing import List, Optional, Dict, Any, Tuple
from core.config import settings
from core.logger import get_logger
from core.utils import get_local_now
import xlsxwriter
from io import BytesIO
from datetime import datetime

logger = get_logger(__name__)


class AssetService(AuditMixin):
    _sort_cache: Dict[str, Dict[str, int]] = {}
    _cache_valid: bool = False

    def __init__(self, db: Session, dict_service: Optional[DictService] = None):
        self.db = db
        self.crypto = CryptoManager(settings.CRYPTO_KEY)
        self.dict_service = dict_service or DictService(db)

    def _refresh_sort_cache(self) -> None:
        if self._cache_valid:
            return
        dict_codes = ["unit", "system", "location", "server_type"]
        for code in dict_codes:
            items = self.dict_service.get_dict_items(code)
            self._sort_cache[code] = {item.item_code: (item.sort_order or 9999) for item in items}
        self._cache_valid = True

    def invalidate_sort_cache(self) -> None:
        self._cache_valid = False
        self._sort_cache.clear()

    def _get_sort_order(self, dict_code: str, item_code: Optional[str]) -> int:
        if not item_code:
            return 9999
        self._refresh_sort_cache()
        return self._sort_cache.get(dict_code, {}).get(item_code, 9999)

    def create(
        self,
        unit_name: str,
        system_name: str,
        username: str,
        password: str,
        ip: Optional[str] = None,
        ipv6: Optional[str] = None,
        port: Optional[int] = None,
        host_name: Optional[str] = None,
        notes: Optional[str] = None,
        business_service: Optional[str] = None,
        location: Optional[str] = None,
        server_type: Optional[str] = None,
        vip: Optional[str] = None,
        user_id: Optional[int] = None
    ) -> Optional[int]:
        if not unit_name or not isinstance(unit_name, str):
            raise ValueError("单位名称不能为空")
        if not system_name or not isinstance(system_name, str):
            raise ValueError("系统名称不能为空")
        if not username or not isinstance(username, str):
            raise ValueError("用户名不能为空")

        password_cipher = self.crypto.encrypt(password) if password else ""

        asset = ServerAsset(
            unit_name=unit_name,
            system_name=system_name,
            ip=ip,
            ipv6=ipv6,
            port=port,
            host_name=host_name,
            username=username,
            password_cipher=password_cipher,
            notes=notes,
            business_service=business_service,
            location=location,
            server_type=server_type,
            vip=vip,
            created_at=get_local_now()
        )
        self.db.add(asset)
        self.db.commit()
        self.db.refresh(asset)

        self.log_audit(
            user_id=user_id,
            action_type="create",
            resource_type="asset",
            resource_id=asset.id,
            details=f"创建资产: {unit_name} - {system_name} ({ip or ipv6})"
        )

        return asset.id

    def get_all(self) -> List[ServerAsset]:
        assets = self.db.query(ServerAsset).all()
        self._refresh_sort_cache()
        assets.sort(key=lambda a: (
            self._sort_cache.get("unit", {}).get(a.unit_name, 9999),
            self._sort_cache.get("system", {}).get(a.system_name, 9999),
            a.ip or "",
            a.ipv6 or ""
        ))
        return assets

    def get_by_id(self, asset_id: int) -> Optional[ServerAsset]:
        if not isinstance(asset_id, int) or asset_id <= 0:
            return None
        return self.db.query(ServerAsset).filter(ServerAsset.id == asset_id).first()

    def update(self, asset_id: int, user_id: Optional[int] = None, **kwargs) -> bool:
        if not isinstance(asset_id, int) or asset_id <= 0:
            return False
        asset = self.get_by_id(asset_id)
        if not asset:
            return False

        for key, value in kwargs.items():
            if hasattr(asset, key) and value is not None:
                if key == "password":
                    setattr(asset, f"{key}_cipher", self.crypto.encrypt(value))
                else:
                    setattr(asset, key, value)

        self.db.commit()

        self.log_audit(
            user_id=user_id,
            action_type="update",
            resource_type="asset",
            resource_id=asset_id,
            details=f"更新资产: {asset.unit_name} - {asset.system_name} ({asset.ip or asset.ipv6})"
        )

        return True

    def delete(self, asset_id: int, user_id: int) -> bool:
        if not isinstance(asset_id, int) or asset_id <= 0:
            return False
        asset = self.get_by_id(asset_id)
        if not asset:
            return False

        self.log_audit(
            user_id=user_id,
            action_type="delete",
            resource_type="asset",
            resource_id=asset_id,
            details=f"删除资产: {asset.unit_name} - {asset.system_name} ({asset.ip or asset.ipv6})"
        )

        self.db.delete(asset)
        self.db.commit()
        return True

    def get_password(self, asset_id: int) -> Optional[str]:
        asset = self.get_by_id(asset_id)
        if not asset or not asset.password_cipher:
            return None
        return self.crypto.decrypt(asset.password_cipher)

    def export_assets(self, asset_ids: Optional[List[int]] = None, include_password: bool = False) -> bytes:
        """
        导出资产数据到Excel文件
        
        Args:
            asset_ids: 要导出的资产ID列表,如果为None则导出所有资产
            include_password: 是否包含密码字段
            
        Returns:
            Excel文件的字节数据
        """
        if asset_ids:
            assets = self.db.query(ServerAsset).filter(ServerAsset.id.in_(asset_ids)).all()
        else:
            assets = self.get_all()
        
        columns = [
            "单位名称", "系统名称", "IP地址", "IPv6地址", "用户名",
            "端口", "主机名", "业务服务", "位置", "服务器类型", "VIP", "备注"
        ]
        if include_password:
            columns.insert(5, "密码")
        
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet('资产列表')
        
        dict_sheet = workbook.add_worksheet('字典数据')
        dict_sheet.hide()
        
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#4472C4',
            'font_color': 'white',
            'border': 1,
            'align': 'center',
            'valign': 'vcenter'
        })
        
        cell_format = workbook.add_format({
            'border': 1,
            'align': 'left',
            'valign': 'vcenter'
        })
        
        for col_num, column in enumerate(columns):
            worksheet.write(0, col_num, column, header_format)
        
        all_data = []
        for row_num, asset in enumerate(assets, start=1):
            unit_name_display = self.dict_service.get_item_name_by_code("unit", asset.unit_name) or asset.unit_name
            system_name_display = self.dict_service.get_item_name_by_code("system", asset.system_name) or asset.system_name
            location_display = self.dict_service.get_item_name_by_code("location", asset.location) or (asset.location or "")
            server_type_display = self.dict_service.get_item_name_by_code("server_type", asset.server_type) or (asset.server_type or "")
            
            data = [
                unit_name_display,
                system_name_display,
                asset.ip or "",
                asset.ipv6 or "",
                asset.username,
                asset.port or 22,
                asset.host_name or "",
                asset.business_service or "",
                location_display,
                server_type_display,
                asset.vip or "",
                asset.notes or ""
            ]
            if include_password:
                data.insert(5, self.crypto.decrypt(asset.password_cipher) if asset.password_cipher else "")
            
            all_data.append(data)
            
            for col_num, value in enumerate(data):
                worksheet.write(row_num, col_num, value, cell_format)
        
        for col_num, column in enumerate(columns):
            max_length = len(str(column))
            for row_data in all_data:
                if col_num < len(row_data):
                    max_length = max(max_length, len(str(row_data[col_num])))
            worksheet.set_column(col_num, col_num, min(max_length + 4, 60))
        
        unit_items = self.dict_service.get_dict_items("unit")
        unit_names = [item.item_name for item in unit_items]
        if unit_names:
            for i, name in enumerate(unit_names):
                dict_sheet.write(i, 0, name)
            max_row = len(unit_names) - 1
            worksheet.data_validation(
                1, 0, max(len(assets), 1), 0,
                {
                    'validate': 'list',
                    'source': f"='字典数据'!$A$1:$A${max_row + 1}",
                    'input_title': '单位名称',
                    'input_message': '请从下拉列表中选择单位名称'
                }
            )
        
        system_items = self.dict_service.get_dict_items("system")
        system_names = [item.item_name for item in system_items]
        if system_names:
            for i, name in enumerate(system_names):
                dict_sheet.write(i, 1, name)
            max_row = len(system_names) - 1
            worksheet.data_validation(
                1, 1, max(len(assets), 1), 1,
                {
                    'validate': 'list',
                    'source': f"='字典数据'!$B$1:$B${max_row + 1}",
                    'input_title': '系统名称',
                    'input_message': '请从下拉列表中选择系统名称'
                }
            )
        
        location_items = self.dict_service.get_dict_items("location")
        location_names = [item.item_name for item in location_items]
        if location_names:
            for i, name in enumerate(location_names):
                dict_sheet.write(i, 2, name)
            max_row = len(location_names) - 1
            location_col = 9 if include_password else 8
            worksheet.data_validation(
                1, location_col, max(len(assets), 1), location_col,
                {
                    'validate': 'list',
                    'source': f"='字典数据'!$C$1:$C${max_row + 1}",
                    'input_title': '位置',
                    'input_message': '请从下拉列表中选择位置'
                }
            )
        
        server_type_items = self.dict_service.get_dict_items("server_type")
        server_type_names = [item.item_name for item in server_type_items]
        if server_type_names:
            for i, name in enumerate(server_type_names):
                dict_sheet.write(i, 3, name)
            max_row = len(server_type_names) - 1
            server_type_col = 10 if include_password else 9
            worksheet.data_validation(
                1, server_type_col, max(len(assets), 1), server_type_col,
                {
                    'validate': 'list',
                    'source': f"='字典数据'!$D$1:$D${max_row + 1}",
                    'input_title': '服务器类型',
                    'input_message': '请从下拉列表中选择服务器类型'
                }
            )
        
        workbook.close()
        output.seek(0)
        
        logger.info(f"导出 {len(assets)} 个资产到Excel文件")
        return output.getvalue()

    def import_assets(
        self, 
        file_data: bytes, 
        update_existing: bool = False,
        skip_errors: bool = True
    ) -> Tuple[int, int, List[str]]:
        """
        从Excel文件导入资产数据
        
        Args:
            file_data: 文件字节数据
            update_existing: 是否更新已存在的资产(根据单位名称+系统名称判断)
            skip_errors: 是否跳过错误继续导入
            
        Returns:
            (成功数量, 失败数量, 错误信息列表)
        """
        try:
            import openpyxl
            workbook = openpyxl.load_workbook(BytesIO(file_data))
            worksheet = workbook.active
            
            rows = []
            for row in worksheet.iter_rows(values_only=True):
                rows.append([cell if cell is not None else "" for cell in row])
            
            if len(rows) == 0:
                return 0, 0, ["文件为空"]
            
            headers = rows[0]
            data_rows = rows[1:]
            
            df = []
            for row in data_rows:
                row_dict = {}
                for i, header in enumerate(headers):
                    if i < len(row):
                        row_dict[header] = row[i]
                    else:
                        row_dict[header] = ""
                df.append(row_dict)
                
        except Exception as e:
            return 0, 0, [f"文件解析失败: {str(e)}"]
        
        success_count = 0
        fail_count = 0
        errors = []
        
        required_fields = ["单位名称", "系统名称", "用户名"]
        
        for index, row_dict in enumerate(df):
            try:
                missing_fields = []
                for field in required_fields:
                    field_with_star = f"{field}*"
                    value = row_dict.get(field) or row_dict.get(field_with_star)
                    if not value or str(value).strip() == "":
                        missing_fields.append(field)
                
                if missing_fields:
                    error_msg = f"第 {index + 2} 行: 缺少必填字段 {', '.join(missing_fields)}"
                    if skip_errors:
                        errors.append(error_msg)
                        fail_count += 1
                        continue
                    else:
                        raise ValueError(error_msg)
                
                unit_name_value = str(row_dict.get("单位名称") or row_dict.get("单位名称*")).strip()
                system_name_value = str(row_dict.get("系统名称") or row_dict.get("系统名称*")).strip()
                username = str(row_dict.get("用户名") or row_dict.get("用户名*")).strip()
                
                unit_name = self.dict_service.get_item_code_by_name("unit", unit_name_value) or unit_name_value
                system_name = self.dict_service.get_item_code_by_name("system", system_name_value) or system_name_value
                
                ip_value = str(row_dict.get("IP地址", "")).strip()
                ipv6_value = str(row_dict.get("IPv6地址", "")).strip()
                
                ip = ip_value if ip_value else None
                ipv6 = ipv6_value if ipv6_value else None
                
                if not ip and not ipv6:
                    error_msg = f"第 {index + 2} 行: IP地址和IPv6地址至少需要填写一个"
                    if skip_errors:
                        errors.append(error_msg)
                        fail_count += 1
                        continue
                    else:
                        raise ValueError(error_msg)
                
                existing_asset = self.db.query(ServerAsset).filter(
                    ServerAsset.unit_name == unit_name,
                    ServerAsset.system_name == system_name,
                    ServerAsset.username == username
                ).filter(
                    (ServerAsset.ip == ip) | (ServerAsset.ip.is_(None) & (ip is None))
                ).filter(
                    (ServerAsset.ipv6 == ipv6) | (ServerAsset.ipv6.is_(None) & (ipv6 is None))
                ).first()
                
                if existing_asset and update_existing:
                    location_value = str(row_dict.get("位置", "")).strip()
                    server_type_value = str(row_dict.get("服务器类型", "")).strip()
                    
                    update_data = {
                        "username": username,
                        "ip": ip,
                        "ipv6": ipv6,
                        "host_name": str(row_dict.get("主机名", "")).strip() or None,
                        "business_service": str(row_dict.get("业务服务", "")).strip() or None,
                        "location": self.dict_service.get_item_code_by_name("location", location_value) or (location_value or None),
                        "server_type": self.dict_service.get_item_code_by_name("server_type", server_type_value) or (server_type_value or None),
                        "vip": str(row_dict.get("VIP", "")).strip() or None,
                        "notes": str(row_dict.get("备注", "")).strip() or None,
                    }
                    
                    port_value = row_dict.get("端口", "22")
                    if port_value and str(port_value).strip():
                        try:
                            update_data["port"] = int(port_value)
                        except ValueError:
                            update_data["port"] = 22
                    else:
                        update_data["port"] = 22
                    
                    password = row_dict.get("密码", "")
                    if password and str(password).strip():
                        update_data["password"] = str(password).strip()
                    
                    for key, value in update_data.items():
                        if key != "password":
                            setattr(existing_asset, key, value)
                        else:
                            setattr(existing_asset, "password_cipher", self.crypto.encrypt(value))
                    
                    self.db.commit()
                    success_count += 1
                    logger.info(f"更新资产: {unit_name} - {system_name}")
                    
                elif not existing_asset:
                    password = row_dict.get("密码", "")
                    if not password:
                        password = ""
                    
                    port_value = row_dict.get("端口", "22")
                    if port_value and str(port_value).strip():
                        try:
                            port = int(port_value)
                        except ValueError:
                            port = 22
                    else:
                        port = 22
                    
                    location_value = str(row_dict.get("位置", "")).strip()
                    server_type_value = str(row_dict.get("服务器类型", "")).strip()
                    
                    asset_id = self.create(
                        unit_name=unit_name,
                        system_name=system_name,
                        username=username,
                        password=str(password).strip(),
                        ip=ip,
                        ipv6=ipv6,
                        port=port,
                        host_name=str(row_dict.get("主机名", "")).strip() or None,
                        business_service=str(row_dict.get("业务服务", "")).strip() or None,
                        location=self.dict_service.get_item_code_by_name("location", location_value) or (location_value or None),
                        server_type=self.dict_service.get_item_code_by_name("server_type", server_type_value) or (server_type_value or None),
                        vip=str(row_dict.get("VIP", "")).strip() or None,
                        notes=str(row_dict.get("备注", "")).strip() or None,
                    )
                    success_count += 1
                    logger.info(f"创建资产: {unit_name} - {system_name}, ID: {asset_id}")
                else:
                    error_msg = f"第 {index + 2} 行: 资产已存在 ({unit_name} - {system_name})"
                    if skip_errors:
                        errors.append(error_msg)
                        fail_count += 1
                        continue
                    else:
                        raise ValueError(error_msg)
                        
            except Exception as e:
                error_msg = f"第 {index + 2} 行: {str(e)}"
                errors.append(error_msg)
                fail_count += 1
                if not skip_errors:
                    self.db.rollback()
                    return success_count, fail_count, errors
        
        logger.info(f"导入完成: 成功 {success_count} 个, 失败 {fail_count} 个")
        return success_count, fail_count, errors
